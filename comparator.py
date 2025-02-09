from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Alignment

def compare_workbooks(file1, file2, compare_style=False):
    wb1 = load_workbook(file1, data_only=False)
    wb2 = load_workbook(file2, data_only=False)

    diff = {}

    for sheet_name in wb1.sheetnames:
        if sheet_name not in wb2.sheetnames:
            diff[sheet_name] = "Missing in second file"
            continue

        ws1 = wb1[sheet_name]
        ws2 = wb2[sheet_name]

        for row in ws1.iter_rows():
            for cell in row:
                ws2_cell = ws2.cell(row=cell.row, column=cell.column)

                cell_diff = {}

                # Compare Values
                if cell.value != ws2_cell.value:
                    cell_diff["Value"] = (cell.value, ws2_cell.value)

                # Compare Formulas
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    if cell.value != ws2_cell.value:
                        cell_diff["Formula"] = (cell.value, ws2_cell.value)

                # Compare Styles
                if compare_style: 
                    if cell.font != ws2_cell.font:
                        cell_diff["Font"] = (cell.font, ws2_cell.font)
                    if cell.fill != ws2_cell.fill:
                        cell_diff["Fill"] = (cell.fill, ws2_cell.fill)
                    if cell.border != ws2_cell.border:
                        cell_diff["Border"] = (cell.border, ws2_cell.border)
                    if cell.alignment != ws2_cell.alignment:
                        cell_diff["Alignment"] = (cell.alignment, ws2_cell.alignment)

                # Compare Hyperlinks
                link1 = ws1.cell(cell.row, cell.column).hyperlink
                link2 = ws2.cell(ws2_cell.row, ws2_cell.column).hyperlink

                if link1 or link2:  # Only compare if at least one has a hyperlink
                    link1_target = link1.target if link1 else None
                    link2_target = link2.target if link2 else None
                    if link1_target != link2_target:
                        cell_diff["Hyperlink"] = (link1_target, link2_target)

                if cell_diff:
                    cell_diff["coordinate"] = cell.coordinate
                    key = (sheet_name, cell.row)
                    if key in diff:
                        diff[key].append(cell_diff)
                    else:
                        diff[key] = [cell_diff]

    return diff


