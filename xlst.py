from openpyxl import load_workbook, Workbook
from copy import copy
from string import Formatter
import re
from jinja2 import Template

from jinja2 import Environment

def concat(*args):
    return "".join(map(str, args))

CELL_ENV = Environment()  # No loader specified
CELL_ENV.globals['CONCAT'] = concat

class ExcelTemplateProcessor:
    def __init__(self, template_path):
        self.template_wb = load_workbook(template_path)
        self.output_wb = load_workbook(template_path)
        #self.output_wb.remove(self.output_wb.active)
        self.formatter = Formatter()

    def process_template(self, context):
        for sheet_name in self.template_wb.sheetnames:
            print("****** Processing sheet", sheet_name)
            if sheet_name.startswith("__"):
                print("skipped")
                continue
            template_ws = self.template_wb[sheet_name]
            output_ws = self.output_wb[sheet_name]
            
            for col in template_ws.column_dimensions:
                output_ws.column_dimensions[col].width = template_ws.column_dimensions[col].width
            
            self._process_sheet(template_ws, output_ws, context)

        if "__post_processing" in self.template_wb.sheetnames:
            print("****** Post processing (hide, fit columns)")
            actions = self._parse_excel_to_objects("__post_processing")
            for action in actions: 
                if action["action"] == "HIDE_COLUMN":
                    print(action)
                    self.output_wb[action["sheet"]].column_dimensions[action["column"]].hidden= True
                elif action["action"] == "FIT_COLUMN":
                    print(action)
                    self._adjust_column_width(self.output_wb[action["sheet"]], action["column"])
                elif action["action"] == "FIT_FIXED_COLUMN":
                    print(action)
                    self.output_wb[action["sheet"]].column_dimensions[action["column"]].width = int(action["arg"])
                elif action["action"] == "ACTIVE_SHEET":
                    print(action)
                    self.output_wb.active = self.output_wb.sheetnames.index(action["sheet"])
                else:                    
                    print("unknown action", action)


    def _parse_excel_to_objects(self,sheet_name):
        sheet = self.template_wb[sheet_name]

        # Extract headers from the first row
        headers = [cell.value for cell in sheet[1]]

        # Parse rows into a list of dictionaries
        data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            obj = dict(zip(headers, row))
            data.append(obj)

        return data


    def _adjust_column_width(self, sheet, column_letter):
        # Find the maximum length of the content in the column
        max_length = 0
        for row in sheet.iter_rows(min_col=sheet[column_letter + '1'].column, max_col=sheet[column_letter + '1'].column):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        
        # Adjust the column width (adding a little padding)
        adjusted_width = (max_length + 2)  # Add 2 for some padding
        sheet.column_dimensions[column_letter].width = adjusted_width


    def _copy_cell_format(self, source_cell, target_cell):
        target_cell.font = copy(source_cell.font)
        target_cell.border = copy(source_cell.border)
        target_cell.fill = copy(source_cell.fill)
        target_cell.number_format = source_cell.number_format
        target_cell.alignment = copy(source_cell.alignment)
        target_cell.protection = copy(source_cell.protection)

    def _process_sheet(self, template_ws, output_ws, context, start_row=1, depth=0, output_row=1):
        row = start_row

        while row <= template_ws.max_row:
            row_values = [str(cell.value) if cell.value else '' for cell in template_ws[row]]
            
            if any('{{range' in val for val in row_values):                            
                range_text = next(val for val in row_values if val and '{{range' in val)
                iterator_name = re.search(r'{{range\s+(\w+)}}', range_text).group(1)
                items = context.get(iterator_name, [])
                
                sub_template_start = row + 1
                sub_template_end, next_row = self._find_matching_end(template_ws, sub_template_start, depth + 1)
                for item_index, item in enumerate(items, 1):
                    print("item", iterator_name, item_index, sub_template_start, sub_template_end, "output_row",output_row)
                    item_context = {
                        **item,
                        "_iterator_name": iterator_name,
                        "_parent_context": context,
                        f"_{iterator_name}_index": item_index,
                        f"_{iterator_name}_count": len(items),
                        "current_row": output_row,
                    }
                    end_row = self._process_sheet(
                        template_ws, 
                        output_ws, 
                        item_context, 
                        start_row=sub_template_start, 
                        depth=depth + 1,
                        output_row=output_row,
                    )
                    output_row = end_row
                    row += end_row
                
                
                row = next_row
                continue
            
            if '{{end}}' in str(row_values):
                return output_row
            
            for col, cell in enumerate(template_ws[row], 1):
                output_cell = output_ws.cell(row=output_row, column=col)
                self._copy_cell_format(cell, output_cell)
                if cell.value and '{{' in str(cell.value):
                    output_cell.value = self._render_template(str(cell.value), context)
                else:
                    output_cell.value = cell.value
            
            output_row += 1
            row += 1
        
        return output_row

    def _find_matching_end(self, ws, start_row, current_depth):
        depth = current_depth
        row = start_row
        
        while row <= ws.max_row:
            row_values = [str(cell.value) if cell.value else '' for cell in ws[row]]
            
            if any('{{range' in val for val in row_values):
                depth += 1
            elif any('{{end}}' in val for val in row_values):
                depth -= 1
                if depth < current_depth:
                    return row - 1, row + 1
            
            row += 1
        
        raise ValueError("No matching {{end}} found for range")

    def _render_template(self, string_template, context):        
        rawtemplate = string_template.replace('“','"').replace('”','"').replace("‘","'").replace("’","'")
        template = CELL_ENV.from_string(rawtemplate)
        rendered = template.render(context)
        # print(template,rendered)
        return rendered

    def save(self, output_path):
        self.output_wb.save(output_path)